# importer.py
import os
import json
import shutil
import secrets
from datetime import datetime
from cryptography.hazmat.primitives.kdf.scrypt import Scrypt
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook

FIXED_COLS = ["Platz", "Wohnung", "Raum", "Typ"]

METER_BLOCK = [
	"Zählernummer",
	"Eingebaut am",
	"Stichtag",
	"Anfangsstand",
	"Endstand",
	"AES-Schlüssel",
	"kc-Faktor",
	"Blockierte Telegramme"
]


def derive_key(password: str, salt: bytes) -> bytes:
	kdf = Scrypt(
		salt=salt,
		length=32,
		n=2**14,
		r=8,
		p=1,
	)
	return kdf.derive(password.encode("utf-8"))


def encrypt(data: bytes, password: str) -> bytes:
	salt = secrets.token_bytes(16)
	nonce = secrets.token_bytes(12)
	key = derive_key(password, salt)
	aesgcm = AESGCM(key)
	ciphertext = aesgcm.encrypt(nonce, data, None)
	return salt + nonce + ciphertext


def read_excel(file_path):
	wb = load_workbook(file_path, data_only=True)
	ws = wb.active

	header = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
	rows = []
	for row in ws.iter_rows(min_row=3):
		rows.append([cell.value for cell in row])
	return header, rows


def import_and_encrypt(excel_path: str, password: str, output_path: str) -> dict:
	if len(password) < 6:
		raise ValueError("Passwort muss mindestens 6 Zeichen lang sein")

	header, rows = read_excel(excel_path)

	fixed_idx = [header.index(col) for col in FIXED_COLS]

	blocks_start = len(FIXED_COLS)
	block_indices = []
	while blocks_start + len(METER_BLOCK) <= len(header):
		block_indices.append(list(range(blocks_start, blocks_start + len(METER_BLOCK))))
		blocks_start += len(METER_BLOCK)

	locations = []
	used_locations = set()
	used_meter_ids = set()

	for row in rows:
		location = str(row[fixed_idx[0]])
		if location in used_locations:
			raise ValueError(f"Platz doppelt: {location}")
		used_locations.add(location)

		entry = {
			"location": location,
			"flat": str(row[fixed_idx[1]]),
			"room": str(row[fixed_idx[2]]),
			"type": str(row[fixed_idx[3]]),
			"meter": []
		}

		for bidx in block_indices:
			raw_id = row[bidx[0]]
			if raw_id is None:
				continue

			zid = str(raw_id).strip().zfill(8)
			if not zid.isdigit() or len(zid) != 8:
				raise ValueError(f"Ungültige Zählernummer: {zid}")
			if zid in used_meter_ids:
				raise ValueError(f"Zähler doppelt: {zid}")
			used_meter_ids.add(zid)

			start_val = row[bidx[1]]
			start_iso = None
			if isinstance(start_val, datetime):
				start_iso = start_val.strftime("%Y-%m-%d")
			elif start_val is not None:
				# ggf. String-Datum direkt übernehmen
				start_iso = str(start_val)

			entry["meter"].append({
				"id": zid,
				"startDate": start_iso,
				"cutoffDate": row[bidx[2]] or "YYYY-01-01",
				"startValue": int(row[bidx[3]]) if row[bidx[3]] else None,
				"finalValue": int(row[bidx[4]]) if row[bidx[4]] else None,
				"aes_key": str(row[bidx[5]]) if row[bidx[5]] else None,
				"kc_factor": float(row[bidx[6]]) if row[bidx[6]] else 1,
				"blockMsg": str(row[bidx[7]]) if row[bidx[7]] else None,
			})

		if not entry["meter"]:
			raise ValueError(f"{location} enthält keinen Zähler")

		locations.append(entry)

	plain = json.dumps({"locations": locations}, indent=2, ensure_ascii=False).encode("utf-8")
	encrypted = encrypt(plain, password)

	if os.path.exists(output_path):
		ts = datetime.now().strftime("%Y%m%d-%H%M%S")
		shutil.copy2(output_path, f"{output_path}.bak-{ts}")

	with open(output_path, "wb") as f:
		f.write(encrypted)

	return {
		"locations": len(locations),
		"meters": sum(len(l["meter"]) for l in locations)
	}
