import sys
import os
import json
import shutil
import secrets
from datetime import datetime
from cryptography.hazmat.primitives.kdf.scrypt import Scrypt
from cryptography.hazmat.primitives.ciphers.aead import AESGCM
from openpyxl import load_workbook

# =========================
# Konfiguration
# =========================

OUTPUT_FILE = "locations.json.enc"

FIXED_COLS = ["Platz", "Wohnung", "Raum", "Typ"]

METER_BLOCK = [
    "Zählernummer",
    "Eingebaut am",
    "Stichtag",
    "Anfangsstand",
    "Endstand",
    "AES-Schlüssel",
    "kc-Faktor",
    "Blockierte Telegramme"
]

# =========================
# Hilfsfunktionen
# =========================

def die(msg):
    print(f"FEHLER: {msg}")
    sys.exit(1)


def backup_file(path: str):
    if not os.path.exists(path):
        return None
    ts = datetime.now().strftime("%Y%m%d-%H%M%S")
    backup_name = f"{path}.bak-{ts}"
    shutil.copy2(path, backup_name)
    return backup_name


def derive_key(password: str, salt: bytes) -> bytes:
    kdf = Scrypt(
        salt=salt,
        length=32,
        n=2**14,
        r=8,
        p=1,
    )
    return kdf.derive(password.encode("utf-8"))


def encrypt(data: bytes, password: str) -> bytes:
    salt = secrets.token_bytes(16)
    nonce = secrets.token_bytes(12)
    key = derive_key(password, salt)
    aesgcm = AESGCM(key)
    ciphertext = aesgcm.encrypt(nonce, data, None)
    return salt + nonce + ciphertext

# =========================
# Excel einlesen
# =========================

def read_excel(file_path):
    wb = load_workbook(file_path, data_only=True)
    ws = wb.active

    # erste Zeile ignorieren, 2. Zeile = Header
    header = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]
    rows = []
    for row in ws.iter_rows(min_row=3):
        rows.append([cell.value for cell in row])
    return header, rows

# =========================
# Main
# =========================

def main():
    if len(sys.argv) != 3:
        die("Aufruf: python import.py locations.xlsx <passwort>")

    excel_file = sys.argv[1]
    password = sys.argv[2]

    if len(password) < 6:
        die("Passwort muss mindestens 6 Zeichen lang sein")
    if not os.path.exists(excel_file):
        die("Excel-Datei nicht gefunden")

    header, rows = read_excel(excel_file)

    # Index der Spalten
    try:
        fixed_idx = [header.index(col) for col in FIXED_COLS]
    except ValueError as e:
        die(f"Fehlende Spalte in Excel: {e}")

    # alle Zählerblöcke erkennen
    blocks_start_idx = len(FIXED_COLS)
    block_indices = []
    while blocks_start_idx + len(METER_BLOCK) <= len(header):
        block_indices.append(list(range(blocks_start_idx, blocks_start_idx + len(METER_BLOCK))))
        blocks_start_idx += len(METER_BLOCK)

    used_locations = set()
    used_meter_ids = set()
    locations = []
    all_starts = []

    for row in rows:
        location = str(row[fixed_idx[0]])
        if location in used_locations:
            die(f"Platz mehrfach vergeben: {location}")
        used_locations.add(location)

        entry = {
            "location": location,
            "flat": str(row[fixed_idx[1]]),
            "room": str(row[fixed_idx[2]]),
            "type": str(row[fixed_idx[3]]),
            "meter": []
        }

        for bidx in block_indices:
            zid_raw = row[bidx[0]]
            if zid_raw is None:
                continue
            # normalize to string, keep digits only
            zid = str(zid_raw).strip()
            if not zid.isdigit():
                die(f"Zähler-ID ist nicht numerisch: {zid}")
            zid = zid.zfill(8) # enforce 8 digits with leading zeros
            if len(zid) != 8:
                die(f"Zähler-ID hat nicht 8 Stellen: {zid}")
            if zid in used_meter_ids:
                die(f"Zähler-ID mehrfach vorhanden: {zid}")
            used_meter_ids.add(zid)

            start_val = row[bidx[1]]
            start_iso = None
            if isinstance(start_val, datetime):
                start_iso = start_val.strftime("%Y-%m-%d")
            elif start_val is not None:
                # ggf. String-Datum direkt übernehmen
                start_iso = str(start_val)
            if start_iso:
                all_starts.append(start_iso)

            entry["meter"].append({
                "id": zid,
                "startDate":  start_iso,
                "cutoffDate":       row[bidx[2]]  if row[bidx[2]] is not None else "YYYY-01-01",
                "startValue": int(  row[bidx[3]]) if row[bidx[3]] is not None else None,
                "finalValue": int(  row[bidx[4]]) if row[bidx[4]] is not None else None,
                "aes_key":    str(  row[bidx[5]]) if row[bidx[5]] is not None else None,
                "kc_factor":  float(row[bidx[6]]) if row[bidx[6]] is not None else 1,
                "blockMsg":   str(  row[bidx[7]]) if row[bidx[7]] is not None else None,
            })

        if not entry["meter"]:
            die(f"Platz {location} enthält keinen Zähler")

        # neuster Zähler zuerst
        entry["meter"].sort(key=lambda z: z["startDate"] or "", reverse=True)
        locations.append(entry)

    # =========================
    # JSON erzeugen
    # =========================
    plain_json = json.dumps(
        {"locations": locations},
        indent=2,
        ensure_ascii=False
    ).encode("utf-8")

    encrypted = encrypt(plain_json, password)

    # =========================
    # Backup & Schreiben
    # =========================
    backup = None
    if os.path.exists(OUTPUT_FILE):
        backup = backup_file(OUTPUT_FILE)

    with open(OUTPUT_FILE, "wb") as f:
        f.write(encrypted)

    # =========================
    # Summary
    # =========================
    print("OK")
    print(f"- {len(locations)} Zählerplätze")
    print(f"- {sum(len(loc['meter']) for loc in locations)} Zähler gesamt")
    if all_starts:
        print(f"- ältester Start: {min(all_starts)}")
        print(f"- neuester Start: {max(all_starts)}")
    print("- keine Konflikte gefunden")
    if backup:
        print(f"- Backup angelegt: {backup}")


if __name__ == "__main__":
    main()
